"""
Data Ingestion Pipeline for Hybrid RAG System
Reads documents, cleans text, chunks, embeds, and stores in PostgreSQL
"""

import os
import sys
import logging
from pathlib import Path
from typing import List, Dict, Any
import re
import json
from datetime import datetime

# Environment and configuration
from dotenv import load_dotenv

# Document processing
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
import openpyxl

# NLP and text processing
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords

# LangChain
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document

# Ollama
import ollama

# Database
import psycopg2
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
from psycopg2.extras import execute_values, Json

# Utils
from tqdm import tqdm
import numpy as np

# Configure logging with UTF-8 encoding for Windows compatibility
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'ingest_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class DataIngestor:
    """Main class for document ingestion pipeline"""

    def __init__(self):
        """Initialize the data ingestor with configuration from .env"""
        logger.info("Initializing DataIngestor...")

        # Load environment variables
        load_dotenv()

        # PostgreSQL configuration
        self.db_config = {
            'dbname': os.getenv('POSTGRES_DB'),
            'user': os.getenv('POSTGRES_USER'),
            'password': os.getenv('POSTGRES_PASSWORD'),
            'host': os.getenv('POSTGRES_HOST', 'localhost'),
            'port': os.getenv('POSTGRES_PORT', '5432')
        }
        self.table_name = os.getenv('VECTOR_TABLE_NAME', 'documents')

        # Embedding configuration
        self.embed_model = os.getenv('OLLAMA_EMBED_MODEL', 'nomic-embed-text:latest')
        self.ollama_base_url = os.getenv('OLLAMA_BASE_URL', 'http://localhost:11434')

        # Chunking configuration
        self.chunk_size = int(os.getenv('CHUNK_SIZE', '1024'))
        self.chunk_overlap = int(os.getenv('CHUNK_OVERLAP', '256'))

        logger.info(f"Database: {self.db_config['dbname']}, Table: {self.table_name}")
        logger.info(f"Chunk size: {self.chunk_size}, Overlap: {self.chunk_overlap}")
        logger.info(f"Embedding model: {self.embed_model}")

        # Download required NLTK data
        self._download_nltk_data()

        # Initialize components
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=self.chunk_size,
            chunk_overlap=self.chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""]
        )

        # Setup database
        self._setup_database()

        logger.info("DataIngestor initialized successfully")

    def _download_nltk_data(self):
        """Download required NLTK data packages"""
        logger.info("Downloading NLTK data packages...")
        try:
            nltk.data.find('tokenizers/punkt')
            nltk.data.find('corpora/stopwords')
        except LookupError:
            nltk.download('punkt', quiet=True)
            nltk.download('stopwords', quiet=True)
            nltk.download('punkt_tab', quiet=True)
        logger.info("NLTK data packages ready")

    def _setup_database(self):
        """Setup PostgreSQL database with pgvector extension"""
        logger.info("Setting up database...")
        try:
            conn = psycopg2.connect(**self.db_config)
            conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
            cur = conn.cursor()

            # Enable pgvector extension
            logger.info("Enabling pgvector extension...")
            cur.execute("CREATE EXTENSION IF NOT EXISTS vector;")

            # Create table with vector column
            logger.info(f"Creating table '{self.table_name}' if not exists...")
            create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {self.table_name} (
                id SERIAL PRIMARY KEY,
                content TEXT NOT NULL,
                metadata JSONB,
                embedding vector(768),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """
            cur.execute(create_table_query)

            # Create index on vector column for faster similarity search
            # Using HNSW index which works on empty tables (unlike IVFFlat)
            logger.info("Creating vector index...")
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS {self.table_name}_embedding_idx
                ON {self.table_name}
                USING hnsw (embedding vector_cosine_ops);
            """)

            cur.close()
            conn.close()
            logger.info("Database setup completed successfully")

        except Exception as e:
            logger.error(f"Database setup failed: {str(e)}")
            raise

    def read_pdf(self, file_path: str) -> str:
        """Read PDF file using PyPDF2 library"""
        logger.info(f"Reading PDF: {file_path}")
        try:
            reader = PdfReader(file_path)
            text_parts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            text = '\n'.join(text_parts)
            logger.info(f"Successfully extracted text from PDF ({len(text)} characters)")
            return text
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {str(e)}")
            raise

    def read_txt(self, file_path: str) -> str:
        """Read TXT file"""
        logger.info(f"Reading TXT: {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            logger.info(f"Successfully read TXT file ({len(text)} characters)")
            return text
        except Exception as e:
            logger.error(f"Error reading TXT {file_path}: {str(e)}")
            raise

    def read_docx(self, file_path: str) -> str:
        """Read DOCX file"""
        logger.info(f"Reading DOCX: {file_path}")
        try:
            doc = DocxDocument(file_path)
            text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            logger.info(f"Successfully read DOCX file ({len(text)} characters)")
            return text
        except Exception as e:
            logger.error(f"Error reading DOCX {file_path}: {str(e)}")
            raise

    def read_xlsx(self, file_path: str) -> str:
        """Read XLSX file"""
        logger.info(f"Reading XLSX: {file_path}")
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        text_parts.append(row_text)

            text = '\n'.join(text_parts)
            logger.info(f"Successfully read XLSX file ({len(text)} characters)")
            return text
        except Exception as e:
            logger.error(f"Error reading XLSX {file_path}: {str(e)}")
            raise

    def read_file(self, file_path: str) -> str:
        """Read file based on extension"""
        file_ext = Path(file_path).suffix.lower()

        if file_ext == '.pdf':
            return self.read_pdf(file_path)
        elif file_ext == '.txt':
            return self.read_txt(file_path)
        elif file_ext == '.docx':
            return self.read_docx(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self.read_xlsx(file_path)
        else:
            logger.warning(f"Unsupported file type: {file_ext}")
            return ""

    def clean_text(self, text: str) -> str:
        """Clean text using NLTK"""
        logger.info("Cleaning text...")
        try:
            # Remove extra whitespace
            text = re.sub(r'\s+', ' ', text)

            # Remove special characters but keep sentence structure
            text = re.sub(r'[^\w\s\.\,\!\?\-\:\;]', '', text)

            # Remove URLs
            text = re.sub(r'http\S+|www\S+', '', text)

            # Remove email addresses
            text = re.sub(r'\S+@\S+', '', text)

            # Normalize whitespace
            text = text.strip()

            logger.info(f"Text cleaned ({len(text)} characters)")
            return text

        except Exception as e:
            logger.error(f"Error cleaning text: {str(e)}")
            return text

    def chunk_text(self, text: str, metadata: Dict[str, Any]) -> List[Document]:
        """Chunk text using LangChain RecursiveCharacterTextSplitter"""
        logger.info(f"Chunking text into segments (size={self.chunk_size}, overlap={self.chunk_overlap})...")
        try:
            documents = self.text_splitter.create_documents([text], [metadata])
            logger.info(f"Created {len(documents)} chunks")
            return documents
        except Exception as e:
            logger.error(f"Error chunking text: {str(e)}")
            raise

    def embed_text(self, text: str) -> List[float]:
        """Generate embeddings using Ollama"""
        try:
            response = ollama.embeddings(
                model=self.embed_model,
                prompt=text
            )
            return response['embedding']
        except Exception as e:
            logger.error(f"Error generating embedding: {str(e)}")
            raise

    def store_chunks(self, chunks: List[Document]):
        """Store chunks with embeddings in PostgreSQL"""
        logger.info(f"Storing {len(chunks)} chunks in database...")

        try:
            conn = psycopg2.connect(**self.db_config)
            cur = conn.cursor()

            # Prepare data for bulk insert
            data_to_insert = []

            for chunk in tqdm(chunks, desc="Generating embeddings"):
                try:
                    # Generate embedding
                    embedding = self.embed_text(chunk.page_content)

                    # Prepare data tuple
                    data_to_insert.append((
                        chunk.page_content,
                        Json(chunk.metadata),
                        embedding
                    ))

                except Exception as e:
                    logger.error(f"Error processing chunk: {str(e)}")
                    continue

            # Bulk insert
            logger.info(f"Inserting {len(data_to_insert)} chunks into database...")
            insert_query = f"""
                INSERT INTO {self.table_name} (content, metadata, embedding)
                VALUES %s
            """

            execute_values(
                cur,
                insert_query,
                data_to_insert,
                template="(%s, %s, %s::vector)"
            )

            conn.commit()
            cur.close()
            conn.close()

            logger.info(f"Successfully stored {len(data_to_insert)} chunks")

        except Exception as e:
            logger.error(f"Error storing chunks: {str(e)}")
            raise

    def process_folder(self, folder_path: str):
        """Process all files in the specified folder"""
        logger.info(f"Processing folder: {folder_path}")

        folder = Path(folder_path)
        if not folder.exists():
            logger.error(f"Folder does not exist: {folder_path}")
            raise FileNotFoundError(f"Folder not found: {folder_path}")

        # Supported file extensions
        supported_extensions = {'.pdf', '.txt', '.docx', '.xlsx', '.xls'}

        # Find all files
        files = [f for f in folder.rglob('*') if f.suffix.lower() in supported_extensions]

        if not files:
            logger.warning(f"No supported files found in {folder_path}")
            return

        logger.info(f"Found {len(files)} files to process")

        # Process each file
        for file_path in files:
            try:
                logger.info(f"\n{'='*80}")
                logger.info(f"Processing file: {file_path.name}")
                logger.info(f"{'='*80}")

                # Read file
                text = self.read_file(str(file_path))

                if not text or len(text.strip()) < 10:
                    logger.warning(f"Skipping file with insufficient content: {file_path.name}")
                    continue

                # Clean text
                cleaned_text = self.clean_text(text)

                # Create metadata
                metadata = {
                    'source': str(file_path),
                    'filename': file_path.name,
                    'file_type': file_path.suffix,
                    'processed_at': datetime.now().isoformat()
                }

                # Chunk text
                chunks = self.chunk_text(cleaned_text, metadata)

                # Store chunks with embeddings
                self.store_chunks(chunks)

                logger.info(f"Successfully processed: {file_path.name}")

            except Exception as e:
                logger.error(f"Failed to process {file_path.name}: {str(e)}")
                continue

        logger.info(f"\n{'='*80}")
        logger.info("Processing completed!")
        logger.info(f"{'='*80}")


def main():
    """Main execution function"""
    if len(sys.argv) < 2:
        print("Usage: python ingest.py <folder_path>")
        print("Example: python ingest.py ./data")
        sys.exit(1)

    folder_path = sys.argv[1]

    try:
        ingestor = DataIngestor()
        ingestor.process_folder(folder_path)
        logger.info("Data ingestion completed successfully!")

    except Exception as e:
        logger.error(f"Data ingestion failed: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
